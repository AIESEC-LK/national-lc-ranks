import requests
import pandas as pd
from openpyxl import Workbook
from datetime import datetime

base_url = 'https://analytics.api.aiesec.org/v2/applications/analyze.json'
access_token = '07b2de66c807bb9873a20e3114060120d34f119eb94bf80efca359ef93ce6c2c'

entities_list = [
    {'id': '222', 'name': 'CC'},
    {'id': '872', 'name': 'CN'},
    {'id': '1340', 'name': 'CS'},
    {'id': '221', 'name': 'USJ'},
    {'id': '2204', 'name': 'Kandy'},
    {'id': '2175', 'name': 'Ruhuna'},
    {'id': '2188', 'name': 'SLIIT'},
    {'id': '2186', 'name': 'NSBM'},
    {'id': '4535', 'name': 'NIBM'},
    {'id': '5490', 'name': 'Rajarata'}
]

funnel_stages = {
    "open": "SU",
    "applied": "APL",
    "an_accepted": "ACC",
    "approved": "APD",
    "realized": "REA",
    "finished": "FIN",
    "completed": "CMP"
}

product_patterns = [
    {"name": "Total", "pattern": ".*_total$"},
    {"name": "oGV", "pattern": "o_.*_7$"},
    {"name": "oGTa", "pattern": "o_.*_8$"},
    {"name": "oGTe", "pattern": "o_.*_9$"},
    {"name": "iGV", "pattern": "i_.*_7$"},
    {"name": "iGTa", "pattern": "i_.*_8$"},
    {"name": "iGTe", "pattern": "i_.*_9$"}
]

def get_start_and_end_date(year, month):
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{pd.Period(start_date).days_in_month}"
    return start_date, end_date

def get_current_year_and_month():
    now = datetime.now()
    return now.year, now.month

def fetch_data_for_entity(start_date, end_date, entity_id):
    url = f"{base_url}?access_token={access_token}&start_date={start_date}&end_date={end_date}&performance_v3%5Boffice_id%5D={entity_id}"
    response = requests.get(url)
    response.raise_for_status()
    return response.json()

def extract_data_by_entity_and_stage(data):
    initial_data = {short_code: {} for short_code in funnel_stages.values()}

    for key, value in data.items():
        for stage, short_code in funnel_stages.items():
            if stage in key:
                if key.startswith('open'):
                    extracted_value = value.get('doc_count', 0)
                else:
                    extracted_value = value.get('applicants', {}).get('value', 0)

                if key in initial_data[short_code]:
                    initial_data[short_code][key] += extracted_value
                else:
                    initial_data[short_code][key] = extracted_value

    return initial_data

def filter_data_by_product_patterns(initial_data):
    final_data = {short_code: {product['name']: 0 for product in product_patterns} for short_code in funnel_stages.values()}

    for stage, keys in initial_data.items():
        for key, value in keys.items():
            for product in product_patterns:
                if pd.Series([key]).str.contains(product['pattern']).any():
                    final_data[stage][product['name']] += value

    return final_data

def create_excel_sheets(final_output):
    wb = Workbook()

    for month, data in final_output.items():
        ws = wb.create_sheet(title=str(month))

        # Headers for stages
        stages = list(funnel_stages.values())
        products = [product['name'] for product in product_patterns]

        # Write headers
        for col, stage in enumerate(stages, start=2):
            ws.merge_cells(start_row=1, start_column=col * 7 - 5, end_row=1, end_column=col * 7 + 1)
            ws.cell(row=1, column=col * 7 - 5).value = stage

            for sub_col, product in enumerate(products, start=0):
                ws.cell(row=2, column=col * 7 - 5 + sub_col).value = product

        # Write entity names
        for row, entity in enumerate(entities_list, start=3):
            ws.cell(row=row, column=1).value = entity['name']

            for col, stage in enumerate(stages, start=2):
                for sub_col, product in enumerate(products, start=0):
                    ws.cell(row=row, column=col * 7 - 5 + sub_col).value = data[entity['name']][stage][product]

    wb.save("AIESEC_Data.xlsx")
    

def main():
    current_year, current_month = get_current_year_and_month()
    date_structure = {month: get_start_and_end_date(current_year, month) for month in range(1, current_month + 1)}

    final_output = {}
    for month, dates in date_structure.items():
        monthly_data = {}
        for entity in entities_list:
            entity_data = fetch_data_for_entity(dates[0], dates[1], entity['id'])
            initial_data = extract_data_by_entity_and_stage(entity_data)
            filtered_data = filter_data_by_product_patterns(initial_data)
            monthly_data[entity['name']] = filtered_data

        final_output[month] = monthly_data

    create_excel_sheets(final_output)

if __name__ == "__main__":
    main()
